import base64
import io
import json
import sqlite3
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path

from flask import Flask, request, jsonify, send_file, render_template, g
from anthropic import Anthropic
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32MB max request size

DB_PATH = Path(__file__).parent / 'bullseye.db'


def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(str(DB_PATH))
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(str(DB_PATH))
    db.execute('''CREATE TABLE IF NOT EXISTS scans (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        shooter_name TEXT,
        match_date TEXT,
        grand_aggregate TEXT,
        score_data TEXT NOT NULL,
        image_b64 TEXT,
        created_at TEXT NOT NULL
    )''')
    db.commit()
    db.close()


init_db()

EXTRACT_PROMPT = """You are analyzing a Bullseye Pistol League score sheet image. Extract ALL data into the following JSON structure.

Scoring rules:
- Each shot scores 0-10. "X" means bullseye (scores as 10 but recorded as "X").
- "M" means miss (scores as 0).
- Each stage (Slow Fire, Timed Fire, Rapid Fire) has 20 shots in two strings of 10.
- Stage total format: "score-Xcount" (e.g., "187-3" means 187 points with 3 Xs).
- Aggregate is sum of all 3 stages for that match type.
- Grand Aggregate = Rimfire Aggregate + Centerfire Aggregate.

Return ONLY valid JSON (no markdown, no code fences) with this structure:
{
  "rimfire": {
    "name": "",
    "date": "",
    "class": "",
    "tgt_no": "",
    "slow_fire": {
      "string1": [null,null,null,null,null,null,null,null,null,null],
      "string2": [null,null,null,null,null,null,null,null,null,null],
      "total": ""
    },
    "timed_fire": {
      "string1": [null,null,null,null,null,null,null,null,null,null],
      "string2": [null,null,null,null,null,null,null,null,null,null],
      "total": ""
    },
    "rapid_fire": {
      "string1": [null,null,null,null,null,null,null,null,null,null],
      "string2": [null,null,null,null,null,null,null,null,null,null],
      "total": ""
    },
    "aggregate": ""
  },
  "centerfire": {
    "name": "",
    "date": "",
    "class": "",
    "tgt_no": "",
    "slow_fire": {
      "string1": [null,null,null,null,null,null,null,null,null,null],
      "string2": [null,null,null,null,null,null,null,null,null,null],
      "total": ""
    },
    "timed_fire": {
      "string1": [null,null,null,null,null,null,null,null,null,null],
      "string2": [null,null,null,null,null,null,null,null,null,null],
      "total": ""
    },
    "rapid_fire": {
      "string1": [null,null,null,null,null,null,null,null,null,null],
      "string2": [null,null,null,null,null,null,null,null,null,null],
      "total": ""
    },
    "aggregate": ""
  },
  "grand_aggregate": ""
}

For each shot, use:
- "X" for bullseye hits
- Integer 0-10 for scored shots
- null for empty/unfilled cells

For totals and aggregates, use the "score-Xcount" format as written (e.g., "187-3"), or "" if empty.

Be very careful reading handwritten numbers. Look closely at each cell in the grid."""


MAX_IMAGE_BYTES = 4 * 1024 * 1024  # 4MB to stay under Claude's 5MB base64 limit


def convert_to_jpeg(file_bytes, filename):
    """Convert uploaded image to JPEG bytes, handling HEIC and other formats.
    Ensures the result is under MAX_IMAGE_BYTES by resizing if needed."""
    from PIL import Image

    suffix = Path(filename).suffix.lower()

    if suffix in ('.heic', '.heif'):
        # Use sips to convert HEIC to JPEG first, then process with Pillow
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp_in:
            tmp_in.write(file_bytes)
            tmp_in.flush()
            tmp_out = tmp_in.name + '.jpg'
            subprocess.run(
                ['sips', '-s', 'format', 'jpeg',
                 tmp_in.name, '--out', tmp_out],
                capture_output=True, check=True
            )
            with open(tmp_out, 'rb') as f:
                file_bytes = f.read()

    # Open with Pillow and compress/resize as needed
    img = Image.open(io.BytesIO(file_bytes))
    img = img.convert('RGB')

    # Try quality levels first, then resize if still too large
    for quality in (85, 70, 55, 40):
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=quality)
        if buf.tell() <= MAX_IMAGE_BYTES:
            return buf.getvalue(), 'image/jpeg'

    # Still too large — resize down
    while True:
        w, h = img.size
        img = img.resize((int(w * 0.75), int(h * 0.75)), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=70)
        if buf.tell() <= MAX_IMAGE_BYTES:
            return buf.getvalue(), 'image/jpeg'


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/extract', methods=['POST'])
def extract():
    api_key = request.headers.get('X-API-Key')
    if not api_key:
        return jsonify({'error': 'API key is required. Set it in Settings.'}), 400

    if 'image' not in request.files:
        return jsonify({'error': 'No image uploaded'}), 400

    file = request.files['image']
    file_bytes = file.read()

    try:
        img_bytes, content_type = convert_to_jpeg(file_bytes, file.filename)
    except Exception as e:
        return jsonify({'error': f'Failed to process image: {e}'}), 400

    img_b64 = base64.b64encode(img_bytes).decode('utf-8')

    try:
        client = Anthropic(api_key=api_key)
        message = client.messages.create(
            model='claude-sonnet-4-20250514',
            max_tokens=4096,
            messages=[{
                'role': 'user',
                'content': [
                    {
                        'type': 'image',
                        'source': {
                            'type': 'base64',
                            'media_type': content_type,
                            'data': img_b64,
                        },
                    },
                    {
                        'type': 'text',
                        'text': EXTRACT_PROMPT,
                    },
                ],
            }],
        )

        response_text = message.content[0].text.strip()
        # Strip markdown code fences if present
        if response_text.startswith('```'):
            response_text = response_text.split('\n', 1)[1]
            response_text = response_text.rsplit('```', 1)[0].strip()

        data = json.loads(response_text)
        # Include the converted JPEG so the browser can preview it (HEIC won't display natively)
        data['_preview_b64'] = img_b64
        return jsonify(data)

    except json.JSONDecodeError:
        return jsonify({'error': 'Failed to parse score data from image. Please try again.'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export', methods=['POST'])
def export():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    # Extract image data before building spreadsheet
    image_b64 = data.pop('_image_b64', None)
    data.pop('_preview_b64', None)

    try:
        return _build_excel(data, image_b64)
    except Exception as e:
        return jsonify({'error': f'Failed to generate Excel: {e}'}), 500


def _build_excel(data, image_b64):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Bullseye Scores'

    # Styles
    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    label_font = Font(bold=True, size=10)
    data_font = Font(size=10)
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    stage_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    total_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

    # Column widths
    for col in range(1, 16):
        ws.column_dimensions[chr(64 + col)].width = 6
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['M'].width = 14

    def calc_row_total(shots):
        """Calculate row subtotal as 'score-Xcount' from a list of shot values."""
        score, xs, has_any = 0, 0, False
        for v in (shots or []):
            if v is None or v == '':
                continue
            s = str(v).strip().upper()
            if s == 'X':
                score += 10; xs += 1; has_any = True
            elif s == 'M':
                has_any = True
            else:
                try:
                    score += int(float(s)); has_any = True
                except (ValueError, TypeError):
                    pass
        if not has_any:
            return ''
        return f'{score}-{xs}' if xs > 0 else str(score)

    def write_match_section(ws, start_row, match_data, match_label, section_num_start):
        r = start_row

        # Match header
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
        cell = ws.cell(row=r, column=1, value=match_label)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
        for c in range(1, 15):
            ws.cell(row=r, column=c).border = border
        r += 1

        # Shooter info row
        info_labels = [
            ('NAME', match_data.get('name', '')),
            ('DATE', match_data.get('date', '')),
            ('CLASS', match_data.get('class', '')),
            ('TGT NO.', match_data.get('tgt_no', '')),
        ]
        col = 1
        for label, value in info_labels:
            cell = ws.cell(row=r, column=col, value=label)
            cell.font = label_font
            cell.border = border
            cell.fill = header_fill
            col += 1
            if label == 'NAME':
                ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 2)
                cell = ws.cell(row=r, column=col, value=value)
                cell.font = data_font
                cell.alignment = left_align
                for cc in range(col, col + 3):
                    ws.cell(row=r, column=cc).border = border
                col += 3
            else:
                cell = ws.cell(row=r, column=col, value=value)
                cell.font = data_font
                cell.alignment = center
                cell.border = border
                col += 1
        r += 1

        # Stages
        stages = [
            (f'{section_num_start}. 20 SHOTS SLOW FIRE', 'slow_fire'),
            (f'{section_num_start + 1}. 20 SHOTS TIMED FIRE', 'timed_fire'),
            (f'{section_num_start + 2}. 20 SHOTS RAPID FIRE', 'rapid_fire'),
        ]

        for stage_label, stage_key in stages:
            stage = match_data.get(stage_key, {})

            # Stage header
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
            cell = ws.cell(row=r, column=1, value=stage_label)
            cell.font = subheader_font
            cell.fill = stage_fill
            for c in range(1, 15):
                ws.cell(row=r, column=c).border = border
            r += 1

            # Column numbers row
            ws.cell(row=r, column=1, value='').border = border
            for i in range(10):
                cell = ws.cell(row=r, column=2 + i, value=i + 1)
                cell.font = label_font
                cell.alignment = center
                cell.border = border
                cell.fill = header_fill
            # TOTAL header for row-level totals
            cell = ws.cell(row=r, column=12, value='TOTAL')
            cell.font = label_font
            cell.alignment = center
            cell.border = border
            cell.fill = header_fill
            # No header for the combined stage total column
            ws.merge_cells(start_row=r, start_column=13, end_row=r, end_column=14)
            for cc in range(13, 15):
                ws.cell(row=r, column=cc).border = border
            r += 1

            # String 1
            stage_name = stage_key.replace('_', ' ').upper().split(' ')[0]
            cell = ws.cell(row=r, column=1, value=stage_name)
            cell.font = label_font
            cell.border = border
            s1 = stage.get('string1', [None] * 10)
            for i in range(10):
                val = s1[i] if i < len(s1) else None
                cell = ws.cell(row=r, column=2 + i, value=val if val is not None else '')
                cell.font = data_font
                cell.alignment = center
                cell.border = border
            # Row total for string 1
            cell = ws.cell(row=r, column=12, value=calc_row_total(s1))
            cell.font = Font(bold=True, size=10)
            cell.alignment = center
            cell.border = border
            cell.fill = total_fill
            r += 1

            # String 2
            cell = ws.cell(row=r, column=1, value=stage_name)
            cell.font = label_font
            cell.border = border
            s2 = stage.get('string2', [None] * 10)
            for i in range(10):
                val = s2[i] if i < len(s2) else None
                cell = ws.cell(row=r, column=2 + i, value=val if val is not None else '')
                cell.font = data_font
                cell.alignment = center
                cell.border = border
            # Row total for string 2
            cell = ws.cell(row=r, column=12, value=calc_row_total(s2))
            cell.font = Font(bold=True, size=10)
            cell.alignment = center
            cell.border = border
            cell.fill = total_fill

            # Combined stage total (merged across both string rows, cols 13-14)
            ws.merge_cells(start_row=r - 1, start_column=13, end_row=r, end_column=14)
            total_cell = ws.cell(row=r - 1, column=13, value=stage.get('total', ''))
            total_cell.font = Font(bold=True, size=12)
            total_cell.alignment = center
            total_cell.fill = total_fill
            for rr in range(r - 1, r + 1):
                for cc in range(13, 15):
                    ws.cell(row=rr, column=cc).border = border
            r += 1

        # Aggregate row
        agg_label_map = {
            '.22 MATCH (RIMFIRE)': 'RIMFIRE AGGREGATE',
            'C.F. MATCH (CENTERFIRE)': 'CENTERFIRE AGGREGATE',
        }
        agg_label = agg_label_map.get(match_label, 'AGGREGATE')
        num = section_num_start + 3
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        cell = ws.cell(row=r, column=1, value=f'{num}. {agg_label}')
        cell.font = subheader_font
        cell.fill = total_fill
        ws.merge_cells(start_row=r, start_column=12, end_row=r, end_column=14)
        cell = ws.cell(row=r, column=12, value=match_data.get('aggregate', ''))
        cell.font = Font(bold=True, size=12)
        cell.alignment = center
        cell.fill = total_fill
        for c in range(1, 15):
            ws.cell(row=r, column=c).border = border
        r += 1

        return r

    # Title
    ws.merge_cells('A1:N1')
    title_cell = ws.cell(row=1, column=1, value='Sunnyvale Rod & Gun Club — Bullseye Pistol League')
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = center

    row = 3
    # Rimfire section
    row = write_match_section(ws, row, data.get('rimfire', {}), '.22 MATCH (RIMFIRE)', 1)
    row += 1

    # Centerfire section
    row = write_match_section(ws, row, data.get('centerfire', {}), 'C.F. MATCH (CENTERFIRE)', 5)
    row += 1

    # Grand Aggregate
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    cell = ws.cell(row=row, column=1, value='10. GRAND AGGREGATE')
    cell.font = Font(bold=True, size=14)
    cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)
    cell = ws.cell(row=row, column=12, value=data.get('grand_aggregate', ''))
    cell.font = Font(bold=True, size=14)
    cell.alignment = center
    cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    for c in range(1, 15):
        ws.cell(row=row, column=c).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Add original image on a second sheet if provided
    if image_b64:
        from PIL import Image as PILImage
        ws_img = wb.create_sheet('Original Score Sheet')
        img_bytes = base64.b64decode(image_b64)
        pil_img = PILImage.open(io.BytesIO(img_bytes))
        pil_img = pil_img.convert('RGB')
        # Resize to fit nicely in the sheet (max ~800px wide)
        max_w = 800
        if pil_img.width > max_w:
            ratio = max_w / pil_img.width
            pil_img = pil_img.resize((max_w, int(pil_img.height * ratio)), PILImage.LANCZOS)
        img_buf = io.BytesIO()
        pil_img.save(img_buf, format='JPEG', quality=85)
        img_buf.seek(0)
        xl_img = XlImage(img_buf)
        ws_img.add_image(xl_img, 'A1')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    shooter_name = (data.get('centerfire', {}).get('name', '')
                    or data.get('rimfire', {}).get('name', '')
                    or 'scores')
    filename = f'bullseye_{shooter_name.replace(" ", "_")}.xlsx'

    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route('/api/scans', methods=['GET'])
def list_scans():
    db = get_db()
    rows = db.execute(
        'SELECT id, shooter_name, match_date, grand_aggregate, created_at '
        'FROM scans ORDER BY created_at DESC'
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/scans', methods=['POST'])
def save_scan():
    data = request.get_json()
    if not data or 'score_data' not in data:
        return jsonify({'error': 'No score data provided'}), 400

    score_data = data['score_data']
    image_b64 = data.get('image_b64')

    shooter_name = (score_data.get('centerfire', {}).get('name', '')
                    or score_data.get('rimfire', {}).get('name', '')
                    or 'Unknown')
    match_date = (score_data.get('centerfire', {}).get('date', '')
                  or score_data.get('rimfire', {}).get('date', '')
                  or '')
    grand_aggregate = score_data.get('grand_aggregate', '')

    # Strip internal fields before saving
    clean = {k: v for k, v in score_data.items() if not k.startswith('_')}

    db = get_db()
    cursor = db.execute(
        'INSERT INTO scans (shooter_name, match_date, grand_aggregate, score_data, image_b64, created_at) '
        'VALUES (?, ?, ?, ?, ?, ?)',
        (shooter_name, match_date, grand_aggregate, json.dumps(clean),
         image_b64, datetime.now().isoformat())
    )
    db.commit()
    return jsonify({'id': cursor.lastrowid})


@app.route('/api/scans/<int:scan_id>', methods=['GET'])
def get_scan(scan_id):
    db = get_db()
    row = db.execute('SELECT * FROM scans WHERE id = ?', (scan_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Scan not found'}), 404
    result = dict(row)
    result['score_data'] = json.loads(result['score_data'])
    return jsonify(result)


@app.route('/api/scans/<int:scan_id>', methods=['DELETE'])
def delete_scan(scan_id):
    db = get_db()
    db.execute('DELETE FROM scans WHERE id = ?', (scan_id,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/scans/<int:scan_id>/export', methods=['GET'])
def export_scan(scan_id):
    db = get_db()
    row = db.execute('SELECT score_data, image_b64 FROM scans WHERE id = ?', (scan_id,)).fetchone()
    if not row:
        return jsonify({'error': 'Scan not found'}), 404
    score_data = json.loads(row['score_data'])
    image_b64 = row['image_b64']
    try:
        return _build_excel(score_data, image_b64)
    except Exception as e:
        return jsonify({'error': f'Failed to generate Excel: {e}'}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5050, use_reloader=False)
